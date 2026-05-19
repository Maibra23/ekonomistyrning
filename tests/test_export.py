"""Tests for Excel export helper."""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook

from utils.export import export_to_excel


def test_export_single_sheet():
    df = pd.DataFrame({"Post": ["Intäkter", "Kostnader"], "Belopp": [1000, 600]})
    data = export_to_excel({"Resultat": df})

    workbook = load_workbook(io.BytesIO(data))
    assert "Resultat" in workbook.sheetnames
    sheet = workbook["Resultat"]
    assert sheet.cell(row=1, column=1).value == "Post"
    assert sheet.cell(row=2, column=1).value == "Intäkter"
    assert sheet.cell(row=2, column=2).value == 1000


def test_export_multiple_sheets():
    df1 = pd.DataFrame({"A": [1]})
    df2 = pd.DataFrame({"B": [2]})
    data = export_to_excel({"Ettan": df1, "Tvåan": df2})

    workbook = load_workbook(io.BytesIO(data))
    assert set(workbook.sheetnames) == {"Ettan", "Tvåan"}


def test_export_sheet_name_sanitized():
    df = pd.DataFrame({"X": [1]})
    long_name = "A" * 50 + "/illegal*chars"
    data = export_to_excel({long_name: df})

    workbook = load_workbook(io.BytesIO(data))
    sheet_name = workbook.sheetnames[0]
    assert len(sheet_name) <= 31
    assert "/" not in sheet_name and "*" not in sheet_name


# ---------------------------------------------------------------------------
# Task 10.9: embedded charts
# ---------------------------------------------------------------------------

def test_export_with_chart_produces_valid_xlsx():
    """A chart spec should be embedded without breaking the workbook."""
    from utils.export import export_to_excel
    from openpyxl import load_workbook
    import io
    import pandas as pd

    df = pd.DataFrame(
        {"Post": ["A", "B", "C"], "Belopp": [100, 200, 300]},
    )
    charts = {
        "Test": [
            {
                "type": "column",
                "title": "Belopp per post",
                "categories": "A2:A4",
                "values": "B2:B4",
                "position": "E2",
                "x_axis_title": "Post",
                "y_axis_title": "Belopp (kr)",
            }
        ]
    }
    raw = export_to_excel({"Test": df}, charts=charts)
    assert isinstance(raw, bytes)

    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Test"]
    # openpyxl exposes embedded charts via the _charts list
    assert len(getattr(ws, "_charts", [])) >= 1


def test_export_no_charts_for_sheet_is_silent():
    """Charts mapping that doesn't target a sheet should not raise."""
    from utils.export import export_to_excel
    import pandas as pd

    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    raw = export_to_excel(
        {"Sheet1": df},
        charts={"OtherSheet": [{"type": "column", "categories": "A2:A3", "values": "B2:B3"}]},
    )
    assert isinstance(raw, bytes) and len(raw) > 100


def test_export_chart_unknown_type_falls_back_to_column():
    from utils.export import export_to_excel
    from openpyxl import load_workbook
    import io
    import pandas as pd

    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    raw = export_to_excel(
        {"Test": df},
        charts={"Test": [{"type": "nonsense", "categories": "A2:A3", "values": "B2:B3", "position": "D2"}]},
    )
    wb = load_workbook(io.BytesIO(raw))
    assert len(getattr(wb["Test"], "_charts", [])) >= 1
